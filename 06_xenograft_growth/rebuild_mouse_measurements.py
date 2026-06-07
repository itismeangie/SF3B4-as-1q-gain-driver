from __future__ import annotations

from datetime import datetime
from pathlib import Path
import re

import numpy as np
import pandas as pd
from openpyxl import load_workbook
from path_helpers import TABLES_DIR, WORKBOOKS_DIR


XLSX_PATH = WORKBOOKS_DIR / 'mouse_documentation_Angelina_090226_B410_Geyer.xlsx'
OUT_CSV = TABLES_DIR / 'mouse_measurements_long.csv'


def parse_date(value):
    if value is None:
        return None
    if isinstance(value, datetime):
        return pd.Timestamp(value.date())
    text = str(value).strip()
    try:
        return pd.to_datetime(text, format='%d.%m.%y')
    except Exception:
        return None


def tumor_to_volume(value):
    if value is None:
        return np.nan
    text = str(value).strip().lower().replace(',', '.')

    # Project rule: "small" is missing, "no" is zero.
    if text == 'small':
        return np.nan
    if text == 'no':
        return 0.0

    match = re.match(r'^(\d+(?:\.\d+)?)\s*[x×]\s*(\d+(?:\.\d+)?)$', text)
    if match:
        a, b = map(float, match.groups())
        length = max(a, b)
        width = min(a, b)
        return (length * (width ** 2)) / 2.0

    try:
        return float(text)
    except Exception:
        return np.nan


def weight_to_float(value):
    try:
        return float(value)
    except Exception:
        return np.nan


def extract_sheet(ws, sheet_name: str) -> pd.DataFrame:
    col_date = {}
    current = None
    for col in range(1, ws.max_column + 1):
        date_value = parse_date(ws.cell(1, col).value)
        if date_value is not None:
            current = date_value
        col_date[col] = current

    rows = []
    has_mouse_id_column = str(ws.cell(1, 2).value).strip().lower() == 'mouse_id'
    for row in range(3, ws.max_row + 1):
        if has_mouse_id_column:
            mouse_id = ws.cell(row, 2).value
            animal_number = ws.cell(row, 8).value
            cell_line = ws.cell(row, 5).value
            mark = ws.cell(row, 6).value
        else:
            mouse_id = None
            animal_number = ws.cell(row, 4).value
            cell_line = ws.cell(row, 1).value
            mark = ws.cell(row, 2).value

        if animal_number is None and mouse_id is None:
            continue

        for col in range(1, ws.max_column + 1):
            metric = ws.cell(2, col).value
            metric_text = str(metric).strip().lower() if metric is not None else ''
            if metric_text not in ('tumor size', 'tumor volume', 'mouse weight'):
                continue
            date_value = col_date.get(col)
            if date_value is None:
                continue
            raw_value = ws.cell(row, col).value
            if raw_value is None:
                continue

            normalized_metric = 'Mouse weight'
            if metric_text in ('tumor size', 'tumor volume'):
                normalized_metric = 'Tumor size'

            rows.append(
                {
                    'sheet': sheet_name,
                    'mouse_id': str(mouse_id) if mouse_id is not None else None,
                    'animal_number': animal_number,
                    'cell_line': str(cell_line) if cell_line is not None else None,
                    'mark': str(mark) if mark is not None else None,
                    'date': pd.Timestamp(date_value),
                    'metric': normalized_metric,
                    'raw_value': raw_value,
                }
            )

    return pd.DataFrame(rows)


def main():
    TABLES_DIR.mkdir(parents=True, exist_ok=True)

    workbook = load_workbook(XLSX_PATH, data_only=False)
    week_sheets = [s for s in workbook.sheetnames if str(s).strip().lower().startswith('week ')]
    if not week_sheets:
        raise RuntimeError('No Week sheets found in workbook.')

    # Sort Week sheets numerically (Week 1, Week 2, ...).
    def week_num(name: str) -> int:
        try:
            return int(str(name).split()[1])
        except Exception:
            return 10**9

    week_sheets = sorted(week_sheets, key=week_num)
    week1 = workbook[week_sheets[0]]

    id_rows = []
    for row in range(3, week1.max_row + 1):
        mouse_id = week1.cell(row, 2).value
        if mouse_id is None:
            continue
        id_rows.append(
            {
                'mouse_id': str(mouse_id),
                'animal_number': week1.cell(row, 8).value,
                'cell_line': str(week1.cell(row, 5).value)
                if week1.cell(row, 5).value is not None
                else None,
                'mark': str(week1.cell(row, 6).value)
                if week1.cell(row, 6).value is not None
                else None,
            }
        )
    id_df = pd.DataFrame(id_rows)

    dox_rows = []
    for week_name in week_sheets[1:]:
        ws = workbook[week_name]
        for row in range(3, ws.max_row + 1):
            animal_number = ws.cell(row, 4).value
            if animal_number is None:
                continue
            dox = ws.cell(row, 5).value
            group = 'unknown'
            if dox is not None:
                sign = str(dox).strip()
                if sign == '+':
                    group = 'dox+'
                elif sign == '-':
                    group = 'dox-'
            dox_rows.append(
                {
                    'animal_number': animal_number,
                    'dox_group': group,
                    'week_name': week_name,
                    'week_num': week_num(week_name),
                }
            )

    dox_df = pd.DataFrame(dox_rows)
    if not dox_df.empty:
        # Keep latest known dox assignment per animal.
        dox_df = (
            dox_df.sort_values(['animal_number', 'week_num'])
            .drop_duplicates(subset=['animal_number'], keep='last')
            [['animal_number', 'dox_group']]
        )
    else:
        dox_df = pd.DataFrame(columns=['animal_number', 'dox_group'])

    meta = id_df.merge(dox_df, on='animal_number', how='left')
    meta['dox_group'] = meta['dox_group'].fillna('unknown')

    raw = pd.concat(
        [extract_sheet(workbook[name], name) for name in week_sheets],
        ignore_index=True,
    )

    merged = raw.merge(
        meta[['animal_number', 'mouse_id', 'cell_line', 'dox_group']],
        on='animal_number',
        how='left',
        suffixes=('', '_meta'),
    )

    merged['mouse_id'] = merged['mouse_id'].fillna(merged['mouse_id_meta'])
    merged['cell_line'] = merged['cell_line'].fillna(merged['cell_line_meta'])
    merged['dox_group'] = merged['dox_group'].fillna('unknown')
    merged = merged.drop(columns=['mouse_id_meta', 'cell_line_meta'], errors='ignore')

    merged['value'] = np.nan
    tumor_mask = merged['metric'].eq('Tumor size')
    weight_mask = merged['metric'].eq('Mouse weight')
    merged.loc[tumor_mask, 'value'] = merged.loc[tumor_mask, 'raw_value'].apply(tumor_to_volume)
    merged.loc[weight_mask, 'value'] = merged.loc[weight_mask, 'raw_value'].apply(weight_to_float)

    final = merged.dropna(subset=['mouse_id', 'date', 'value']).sort_values(
        ['metric', 'cell_line', 'animal_number', 'date']
    )
    final.to_csv(OUT_CSV, index=False)

    print(f'Wrote {OUT_CSV}')
    print(f'Rows: {len(final)}; mice: {final["mouse_id"].nunique()}')
    print('Tumor dates:', sorted(final[final['metric'].eq('Tumor size')]['date'].dt.strftime('%Y-%m-%d').unique()))


if __name__ == '__main__':
    main()
